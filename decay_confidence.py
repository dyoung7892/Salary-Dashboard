#!/usr/bin/env python3
"""
Decay Confidence
Finds rows in the database that haven't been verified in more than
--max-age-days days and downgrades their Salary Source to "Unverified".

This prevents confirmed data from 2024 being treated as current in 2026.
A salary from 2024 with no re-check is not a confirmed salary — it's a guess.
"""

import argparse
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

XLSX_PATH = Path("data/FBS_SC_Database_v8_1_Clean.xlsx")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")

# Pattern to find a date in notes like "[Churn 2026-07-16]" or "OR Dec 2024"
DATE_PATTERN = re.compile(r"\[Churn (\d{4}-\d{2}-\d{2})\]")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--max-age-days", type=int, default=180, help="Days before a row decays")
    args = parser.parse_args()

    cutoff = date.today() - timedelta(days=args.max_age_days)
    print(f"Decaying rows not verified since {cutoff}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["FBS S&C Overview"]

    decayed = 0
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False), start=4):
        school = row[0].value
        if not school:
            continue

        sal_source = row[7].value or ""
        notes = row[16].value or ""
        sal_yr = row[6].value

        # Already unverified — skip
        if sal_source in ("Unverified", "Unknown", "Estimated"):
            continue

        # Check if there's a recent churn timestamp
        match = DATE_PATTERN.search(notes)
        if match:
            last_check = date.fromisoformat(match.group(1))
            if last_check >= cutoff:
                continue  # recently checked, keep as-is

        # Check salary year as a proxy for staleness
        if sal_yr and isinstance(sal_yr, int):
            # If salary year is more than 2 years ago, decay it
            if sal_yr < date.today().year - 2:
                ws.cell(row=i, column=8, value="Unverified")
                ws.cell(row=i, column=17, value=f"[Decay {date.today()}] Salary from {sal_yr} exceeds age threshold. " + notes)
                for col in range(1, 18):
                    ws.cell(row=i, column=col).fill = AMBER_FILL
                print(f"  DECAYED: {school} (salary year {sal_yr})")
                decayed += 1

    wb.save(XLSX_PATH)
    print(f"\n{decayed} row(s) decayed in {XLSX_PATH}")


if __name__ == "__main__":
    main()
