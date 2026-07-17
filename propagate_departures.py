#!/usr/bin/env python3
"""
Propagate Departures
Every hire is ALSO a departure. This script reads the churn JSON and marks
vacated rows as INVALIDATED so stale data doesn't accumulate silently.

This is the fix for the Michigan/Utah/Colorado State failure mode:
those schools sat with wrong names for 2 cycles because arrivals were
recorded but the holes they left behind were never flagged.
"""

import argparse
import json
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font

XLSX_PATH = Path("data/FBS_SC_Database_v8_1_Clean.xlsx")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
BOLD = Font(bold=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--in", dest="infile", required=True, help="Churn JSON from churn_sentinel.py")
    args = parser.parse_args()

    with open(args.infile) as f:
        churn = json.load(f)

    changes = churn.get("changes", [])
    if not changes:
        print("No changes to propagate.")
        return

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["FBS S&C Overview"]

    for change in changes:
        school = change["school"]
        old_name = change["old_name"]
        new_name = change["new_name"]
        row_idx = change["row_idx"]
        today = date.today().isoformat()

        # Update the coach name
        ws.cell(row=row_idx, column=4, value=new_name)

        # Mark source as needing verification
        ws.cell(row=row_idx, column=8, value="Unverified")

        # Clear old salary — a new hire's salary is unknown until confirmed
        ws.cell(row=row_idx, column=6, value=None)
        ws.cell(row=row_idx, column=7, value=None)

        # Write the update note
        existing_note = ws.cell(row=row_idx, column=17).value or ""
        note = f"[Churn {today}] Name changed: {old_name} → {new_name}. Salary cleared — requires verification."
        ws.cell(row=row_idx, column=17, value=note)

        # Red fill on the row to flag it visually
        for col in range(1, 18):
            ws.cell(row=row_idx, column=col).fill = RED_FILL

        print(f"INVALIDATED row {row_idx}: {school} — {old_name} → {new_name}")

    wb.save(XLSX_PATH)
    print(f"\n{len(changes)} row(s) updated in {XLSX_PATH}")


if __name__ == "__main__":
    main()
