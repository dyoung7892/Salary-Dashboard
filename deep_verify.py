#!/usr/bin/env python3
"""
Deep Verify — Track B
Runs the weekly salary verification slice defined in config/rotation.yml.
Uses SerpAPI to search for salary data for each school in the current slice.

Rotation: week1=P4 public strong states, week2=P4 weak/FOIA, week3=privates,
          week4=G5 MW+AAC, week5=G5 SunBelt+MAC+CUSA, week6=QA only (no search)

Hard rule: will NOT write a salary to a row whose name is flagged as unverified.
"""

import argparse
import json
import os
import re
import time
from datetime import date
from pathlib import Path

import requests
import openpyxl
import yaml

XLSX_PATH = Path("data/FBS_SC_Database_v8_1_Clean.xlsx")
CONFIG_PATH = Path("config/rotation.yml")

SALARY_PATTERNS = [
    re.compile(r'\$\s*([\d,]+)(?:\s*(?:thousand|K|million|M))?', re.IGNORECASE),
    re.compile(r'([\d,]+)\s*(?:per year|annually|salary|contract)', re.IGNORECASE),
    re.compile(r'earn(?:s|ing|ed)?\s+\$\s*([\d,]+)', re.IGNORECASE),
    re.compile(r'paid?\s+\$\s*([\d,]+)', re.IGNORECASE),
]

UNVERIFIED_SOURCES = {"Estimated", "Unverified", "Unknown"}

# Sources we trust as salary confirmation
TRUSTED_SOURCES = [
    "usatoday", "on3.com", "247sports", "the-advocate", "al.com",
    "secsalaries", "texascollegesalaries", "statesalary",
    "herald-leader", "knoxnews", "tallahassee.com",
    "footballscoop", "collegefootball", "athletic.net",
    "tennessean", "gatorzone", "georgiadogs",
]


def load_config():
    with open(CONFIG_PATH) as f:
        return yaml.safe_load(f)


def load_database():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["FBS S&C Overview"]
    db = {}
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True), start=4):
        if not row[0]:
            continue
        db[row[0]] = {
            "row_idx": i,
            "coach": row[3] or "Unknown",
            "conf": row[1],
            "salary": row[5],
            "sal_source": row[7] or "Unknown",
            "private": row[2],
        }
    return db


def parse_salary(text):
    """Extract the most plausible salary figure from a text snippet."""
    for pattern in SALARY_PATTERNS:
        match = pattern.search(text)
        if match:
            raw = match.group(1).replace(",", "")
            try:
                val = int(raw)
                # Sanity check: S&C salaries are $80K–$1.5M
                if 80_000 <= val <= 1_500_000:
                    return val
                # Handle "850" meaning $850K
                if 80 <= val <= 1500:
                    return val * 1000
            except ValueError:
                pass
    return None


def is_trusted_source(url):
    url_lower = url.lower()
    return any(t in url_lower for t in TRUSTED_SOURCES)


def search_salary(school, coach, api_key):
    """Query SerpAPI for salary information."""
    if not api_key:
        return None, None

    query = f'"{coach}" "{school}" football strength conditioning salary 2026 OR 2025'
    try:
        resp = requests.get(
            "https://serpapi.com/search",
            params={"q": query, "api_key": api_key, "num": 5, "engine": "google"},
            timeout=15
        )
        data = resp.json()

        for result in data.get("organic_results", []):
            url = result.get("link", "")
            snippet = result.get("snippet", "")
            title = result.get("title", "")
            full_text = f"{title} {snippet}"

            salary = parse_salary(full_text)
            if salary and is_trusted_source(url):
                return salary, url

        # Second pass: any result with a salary even if not trusted source
        for result in data.get("organic_results", []):
            url = result.get("link", "")
            snippet = result.get("snippet", "")
            salary = parse_salary(snippet)
            if salary:
                return salary, url

    except Exception as e:
        print(f"  Search error for {school}: {e}")

    return None, None


def run_week6_qa(db, ws):
    """Week 6: reconciliation only — no network calls."""
    print("Week 6: QA reconciliation pass")

    # Detect duplicate names (catches the Feeley Duke/UAB misfile class)
    name_to_schools = {}
    for school, row in db.items():
        coach = row["coach"]
        if coach and coach != "Unknown":
            name_to_schools.setdefault(coach, []).append(school)

    duplicates = {k: v for k, v in name_to_schools.items() if len(v) > 1}
    if duplicates:
        print(f"\nDUPLICATE NAMES DETECTED (possible row misfile):")
        for name, schools in duplicates.items():
            print(f"  '{name}' appears in: {', '.join(schools)}")
    else:
        print("  No duplicate names found.")

    return {"duplicates": duplicates, "date": str(date.today())}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--slice", required=True, help="week1-week6")
    parser.add_argument("--config", default="config/rotation.yml")
    parser.add_argument("--out", required=True, help="Output JSON path")
    args = parser.parse_args()

    api_key = os.environ.get("SEARCH_API_KEY", "")
    config = load_config()
    db = load_database()

    slice_cfg = config["slices"].get(args.slice)
    if not slice_cfg:
        print(f"Unknown slice: {args.slice}")
        return

    print(f"\nTrack B — {slice_cfg['label']}")
    print(f"Method: {slice_cfg['method']}\n")

    # Week 6 is internal reconciliation only
    if slice_cfg["method"] == "internal":
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb["FBS S&C Overview"]
        result = run_week6_qa(db, ws)
        wb.save(XLSX_PATH)
        Path(args.out).parent.mkdir(parents=True, exist_ok=True)
        with open(args.out, "w") as f:
            json.dump(result, f, indent=2)
        return

    # Determine which schools to process for this slice
    target_confs = set(slice_cfg.get("conferences", []))
    exclude_private = slice_cfg.get("exclude_private", False)
    specific_schools = slice_cfg.get("schools", [])

    if specific_schools:
        targets = [(s, db[s]) for s in specific_schools if s in db]
    else:
        targets = [
            (school, row) for school, row in db.items()
            if (not target_confs or row["conf"] in target_confs)
            and (not exclude_private or not row["private"])
        ]

    print(f"Processing {len(targets)} schools\n")

    results = {
        "slice": args.slice,
        "date": str(date.today()),
        "confirmed": [],
        "not_found": [],
        "skipped_name_gate": [],
    }

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["FBS S&C Overview"]
    updated = 0

    for school, row in targets:
        coach = row["coach"]
        sal_source = row["sal_source"]
        row_idx = row["row_idx"]

        # NAME GATE: skip if name is unverified
        if coach == "Unknown" or sal_source in UNVERIFIED_SOURCES:
            print(f"GATE: {school} — name unverified, skipping salary search")
            results["skipped_name_gate"].append({"school": school, "coach": coach})
            continue

        print(f"Searching: {school} ({coach})...", end=" ")
        salary, source_url = search_salary(school, coach, api_key)
        time.sleep(2)  # polite delay

        if salary:
            print(f"Found ${salary:,} ({source_url[:60] if source_url else 'no url'!r})")
            # Write back to xlsx
            ws.cell(row=row_idx, column=6, value=salary)
            ws.cell(row=row_idx, column=7, value=date.today().year)
            ws.cell(row=row_idx, column=8, value="Confirmed")
            ws.cell(row=row_idx, column=17, value=f"[DeepVerify {date.today()}] ${salary:,} via {source_url}")
            updated += 1
            results["confirmed"].append({
                "school": school, "coach": coach,
                "salary": salary, "source": source_url
            })
        else:
            print("Not found")
            results["not_found"].append({"school": school, "coach": coach})

    wb.save(XLSX_PATH)
    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    with open(args.out, "w") as f:
        json.dump(results, f, indent=2)

    print(f"\nSlice complete. Confirmed: {updated} | Not found: {len(results['not_found'])} | Gated: {len(results['skipped_name_gate'])}")


if __name__ == "__main__":
    main()
