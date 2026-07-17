#!/usr/bin/env python3
"""
Enforce Name Gate
The single rule that would have prevented every fabrication in v8.1:
a row whose name is not verified cannot receive a salary value.

Scans the xlsx and exits non-zero if any row has a salary value but
an unverified/unknown/estimated name source. Used as a CI gate before
the deep verify step runs.
"""

import argparse
import sys
from pathlib import Path

import openpyxl

XLSX_PATH = Path("data/FBS_SC_Database_v8_1_Clean.xlsx")

UNVERIFIED_SOURCES = {"Estimated", "Unverified", "Unknown"}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--strict", action="store_true", help="Exit 1 if violations found")
    parser.add_argument("--fix", action="store_true", help="Auto-clear violating salaries")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["FBS S&C Overview"]

    violations = []
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False), start=4):
        school = row[0].value
        if not school:
            continue

        coach = row[3].value or "Unknown"
        sal_source = row[7].value or "Unknown"
        salary = row[5].value

        if salary and sal_source in UNVERIFIED_SOURCES:
            violations.append({
                "row": i,
                "school": school,
                "coach": coach,
                "salary": salary,
                "source": sal_source,
            })

            if args.fix:
                ws.cell(row=i, column=6, value=None)   # clear salary
                ws.cell(row=i, column=7, value=None)   # clear year
                print(f"  FIXED: {school} row {i} — salary cleared (source was '{sal_source}')")

    if args.fix and violations:
        wb.save(XLSX_PATH)

    if violations:
        print(f"\nName gate violations: {len(violations)}")
        for v in violations:
            print(f"  Row {v['row']}: {v['school']} | {v['coach']} | ${v['salary']:,} | source='{v['source']}'")
        if args.strict and not args.fix:
            print("\nBlocking: salary data present on unverified rows. Run with --fix to clear them.")
            sys.exit(1)
    else:
        print("Name gate: OK — no salary on unverified rows.")


if __name__ == "__main__":
    main()
