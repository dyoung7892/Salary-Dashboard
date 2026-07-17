#!/usr/bin/env python3
"""
Dashboard Builder
Reads the xlsx and generates docs/index.html — a dark-theme interactive
dashboard showing only confirmed data. Never shows estimated rows as fact.
"""

import argparse
import json
from datetime import date
from pathlib import Path

import openpyxl

XLSX_PATH = Path("data/FBS_SC_Database_v8_1_Clean.xlsx")

CONF_ORDER = ["SEC", "Big Ten", "Big 12", "ACC", "AAC", "Sun Belt", "MAC", "MW", "CUSA"]


def load_rows():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["FBS S&C Overview"]
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "school": row[0],
            "conf": row[1] or "",
            "private": "Yes" if row[2] else "No",
            "coach": row[3] or "Unknown",
            "title": row[4] or "",
            "salary": row[5],
            "sal_yr": row[6],
            "sal_src": row[7] or "Unknown",
            "tier": row[8] or "",
            "paid_asst": row[9] or 0,
            "sports_sci": row[10] or 0,
            "gas": row[11] or 0,
            "interns": row[12] or 0,
            "total_paid": row[13] or 0,
            "staff_src": row[14] or "",
            "notes": row[15] or "",
        })
    return rows


def fmt_salary(val):
    if not val:
        return "Unknown"
    return f"${val:,.0f}"


def build_html(rows, confirmed_only):
    if confirmed_only:
        display_rows = [r for r in rows if r["sal_src"] in ("Confirmed", "Confirmed (name)")]
    else:
        display_rows = rows

    today = date.today().isoformat()
    total = len(rows)
    confirmed_count = len([r for r in rows if r["sal_src"] in ("Confirmed", "Confirmed (name)")])

    rows_json = json.dumps(display_rows)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>FBS S&amp;C Salary Database</title>
<style>
  :root {{
    --bg: #0d0e10; --surface: #16181c; --surface2: #1e2026;
    --border: #2a2d35; --text: #e8e8e8; --muted: #888;
    --crimson: #c8102e; --crimson-light: #ee2233;
    --green: #2da44e; --amber: #f0c940; --red: #e24b4a;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: var(--bg); color: var(--text); font-family: Arial, sans-serif; font-size: 14px; }}
  header {{ background: var(--crimson); padding: 16px 24px; display: flex; align-items: center; gap: 16px; }}
  header h1 {{ font-size: 20px; font-weight: 700; letter-spacing: .5px; }}
  header .sub {{ font-size: 13px; opacity: .8; margin-top: 2px; }}
  .stats {{ display: flex; gap: 16px; padding: 16px 24px; background: var(--surface); border-bottom: 1px solid var(--border); flex-wrap: wrap; }}
  .stat {{ background: var(--surface2); border: 1px solid var(--border); border-radius: 8px; padding: 10px 16px; min-width: 130px; }}
  .stat .val {{ font-size: 22px; font-weight: 700; color: var(--crimson-light); }}
  .stat .lbl {{ font-size: 11px; color: var(--muted); margin-top: 2px; text-transform: uppercase; letter-spacing: .5px; }}
  .controls {{ padding: 12px 24px; background: var(--surface); border-bottom: 1px solid var(--border); display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
  select, input {{ background: var(--surface2); border: 1px solid var(--border); color: var(--text); padding: 6px 10px; border-radius: 6px; font-size: 13px; }}
  .table-wrap {{ overflow-x: auto; padding: 0 24px 24px; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
  th {{ background: var(--surface2); color: var(--muted); font-size: 11px; text-transform: uppercase; letter-spacing: .5px; padding: 8px 12px; text-align: left; border-bottom: 2px solid var(--border); position: sticky; top: 0; cursor: pointer; white-space: nowrap; }}
  th:hover {{ color: var(--text); }}
  td {{ padding: 9px 12px; border-bottom: 1px solid var(--border); vertical-align: top; }}
  tr:hover td {{ background: var(--surface2); }}
  .badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
  .badge-confirmed {{ background: #0e3320; color: var(--green); }}
  .badge-estimated {{ background: #3a2e00; color: var(--amber); }}
  .badge-unverified {{ background: #3a0e0e; color: var(--red); }}
  .badge-unknown {{ background: #222; color: var(--muted); }}
  .salary {{ font-weight: 600; font-family: monospace; }}
  .conf {{ font-size: 11px; color: var(--muted); }}
  #count {{ font-size: 13px; color: var(--muted); }}
  .sort-arrow {{ opacity: .4; }}
  .sort-active {{ opacity: 1; color: var(--crimson-light); }}
</style>
</head>
<body>
<header>
  <div>
    <h1>FBS Football S&amp;C Salary Database</h1>
    <div class="sub">Updated {today} &nbsp;·&nbsp; {confirmed_count} confirmed / {total} total programs</div>
  </div>
</header>

<div class="stats">
  <div class="stat"><div class="val" id="stat-total">—</div><div class="lbl">Showing</div></div>
  <div class="stat"><div class="val" id="stat-confirmed">—</div><div class="lbl">Confirmed salary</div></div>
  <div class="stat"><div class="val" id="stat-avg">—</div><div class="lbl">Avg confirmed salary</div></div>
  <div class="stat"><div class="val" id="stat-max">—</div><div class="lbl">Highest confirmed</div></div>
</div>

<div class="controls">
  <select id="conf-filter"><option value="">All conferences</option></select>
  <select id="src-filter">
    <option value="">All data quality</option>
    <option value="Confirmed">Confirmed only</option>
    <option value="Estimated">Estimated</option>
    <option value="Unverified">Unverified</option>
  </select>
  <input id="search" type="text" placeholder="Search school or coach..." style="min-width:200px">
  <span id="count"></span>
</div>

<div class="table-wrap">
<table id="main-table">
<thead>
<tr>
  <th onclick="sortBy('school')">School <span class="sort-arrow" id="sort-school">↕</span></th>
  <th onclick="sortBy('conf')">Conf</th>
  <th onclick="sortBy('coach')">Head S&amp;C Coach <span class="sort-arrow" id="sort-coach">↕</span></th>
  <th onclick="sortBy('salary')">Salary <span class="sort-arrow" id="sort-salary">↕</span></th>
  <th>Quality</th>
  <th>Total Staff</th>
  <th>Title</th>
</tr>
</thead>
<tbody id="tbody"></tbody>
</table>
</div>

<script>
const ALL_ROWS = {rows_json};
const CONFS = {json.dumps(CONF_ORDER)};

let currentSort = null;
let sortAsc = true;
let filtered = [...ALL_ROWS];

// Populate conf filter
const confSel = document.getElementById('conf-filter');
CONFS.forEach(c => {{ const o = document.createElement('option'); o.value = c; o.textContent = c; confSel.appendChild(o); }});

function badge(src) {{
  if (!src || src === 'Unknown') return '<span class="badge badge-unknown">Unknown</span>';
  if (src.includes('Confirmed')) return '<span class="badge badge-confirmed">Confirmed</span>';
  if (src === 'Estimated') return '<span class="badge badge-estimated">Estimated</span>';
  return '<span class="badge badge-unverified">Unverified</span>';
}}

function fmtSalary(v) {{
  if (!v) return '<span style="color:var(--muted)">—</span>';
  return '<span class="salary">$' + Number(v).toLocaleString() + '</span>';
}}

function render() {{
  const conf = document.getElementById('conf-filter').value;
  const src = document.getElementById('src-filter').value;
  const q = document.getElementById('search').value.toLowerCase();

  filtered = ALL_ROWS.filter(r =>
    (!conf || r.conf === conf) &&
    (!src || (r.sal_src || '').includes(src)) &&
    (!q || r.school.toLowerCase().includes(q) || (r.coach || '').toLowerCase().includes(q))
  );

  if (currentSort) {{
    filtered.sort((a, b) => {{
      let av = a[currentSort], bv = b[currentSort];
      if (typeof av === 'number' && typeof bv === 'number') return sortAsc ? av - bv : bv - av;
      av = (av || '').toString().toLowerCase();
      bv = (bv || '').toString().toLowerCase();
      return sortAsc ? av.localeCompare(bv) : bv.localeCompare(av);
    }});
  }}

  const tbody = document.getElementById('tbody');
  tbody.innerHTML = filtered.map(r => `
    <tr>
      <td><strong>${{r.school}}</strong></td>
      <td><span class="conf">${{r.conf}}</span></td>
      <td>${{r.coach || '<span style="color:var(--muted)">Unknown</span>'}}</td>
      <td>${{fmtSalary(r.salary)}}</td>
      <td>${{badge(r.sal_src)}}</td>
      <td style="text-align:center">${{r.total_paid || '—'}}</td>
      <td style="font-size:12px;color:var(--muted)">${{r.title || ''}}</td>
    </tr>
  `).join('');

  document.getElementById('count').textContent = filtered.length + ' programs';

  const confirmed = filtered.filter(r => r.sal_src && r.sal_src.includes('Confirmed'));
  const salaries = confirmed.map(r => r.salary).filter(Boolean);
  document.getElementById('stat-total').textContent = filtered.length;
  document.getElementById('stat-confirmed').textContent = salaries.length;
  document.getElementById('stat-avg').textContent = salaries.length
    ? '$' + Math.round(salaries.reduce((a,b)=>a+b,0)/salaries.length).toLocaleString()
    : '—';
  document.getElementById('stat-max').textContent = salaries.length
    ? '$' + Math.max(...salaries).toLocaleString()
    : '—';
}}

function sortBy(col) {{
  if (currentSort === col) sortAsc = !sortAsc;
  else {{ currentSort = col; sortAsc = true; }}
  document.querySelectorAll('.sort-arrow').forEach(el => el.className = 'sort-arrow');
  const el = document.getElementById('sort-' + col);
  if (el) {{ el.className = 'sort-arrow sort-active'; el.textContent = sortAsc ? '↑' : '↓'; }}
  render();
}}

document.getElementById('conf-filter').addEventListener('change', render);
document.getElementById('src-filter').addEventListener('change', render);
document.getElementById('search').addEventListener('input', render);
render();
</script>
</body>
</html>"""
    return html


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--confirmed-only", action="store_true")
    parser.add_argument("--out", default="docs/index.html")
    args = parser.parse_args()

    rows = load_rows()
    html = build_html(rows, args.confirmed_only)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html, encoding="utf-8")
    print(f"Dashboard written to {out} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
